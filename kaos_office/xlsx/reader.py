"""XLSX reader — parse Excel workbooks into TabularDocument.

Uses python-calamine (Rust, MIT, 7-28x faster than openpyxl) for data
extraction. Calamine returns Python-native types (int, float, str, bool,
date, datetime, time, timedelta) which map directly to ColumnType.

Entry point: ``parse_xlsx(path) → TabularDocument``

Usage::

    from kaos_office.xlsx.reader import parse_xlsx

    doc = parse_xlsx("report.xlsx")
    for table in doc.tables:
        print(f"{table.name}: {table.row_count} rows x {len(table.columns)} cols")

    # Specific sheets
    doc = parse_xlsx("report.xlsx", sheets=["Revenue", "Expenses"])

    # With formula preservation (slower, requires openpyxl)
    doc = parse_xlsx("report.xlsx", include_formulas=True)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from kaos_content.model.attr import Provenance, SourceRef
from kaos_content.model.metadata import DocumentMetadata
from kaos_content.model.tabular import (
    Column,
    Table,
    TabularDocument,
    infer_column_type,
)

# Calamine is required for XLSX reading
try:
    from python_calamine import CalamineWorkbook
except ImportError as exc:
    msg = (
        "python-calamine is required for XLSX extraction. "
        "Install with: pip install kaos-office[xlsx]"
    )
    raise ImportError(msg) from exc


def parse_xlsx(
    path: str | Path,
    *,
    sheets: list[str] | None = None,
    max_rows: int | None = None,
    header_row: int = 0,
    include_formulas: bool = False,
) -> TabularDocument:
    """Parse an XLSX file into a TabularDocument.

    Uses python-calamine (Rust) for fast extraction. Each worksheet
    becomes a Table in the resulting TabularDocument.

    Args:
        path: Path to the XLSX file.
        sheets: Specific sheet names to extract. None = all visible sheets.
        max_rows: Maximum data rows per sheet. None = all rows.
        header_row: Row index (0-based) to use as column headers.
            Rows before the header are skipped.
        include_formulas: If True, extract formulas via openpyxl and
            store in ``table.metadata["formulas"]``. Slower.

    Returns:
        TabularDocument with one Table per extracted sheet.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not a valid XLSX.
    """
    p = Path(path).resolve()
    if not p.is_file():
        msg = f"File not found: {p}"
        raise FileNotFoundError(msg)

    wb = CalamineWorkbook.from_path(str(p))

    # Determine which sheets to extract
    if sheets is not None:
        sheet_names = sheets
    else:
        # All visible worksheets
        from python_calamine import SheetTypeEnum, SheetVisibleEnum

        sheet_names = [
            sm.name
            for sm in wb.sheets_metadata
            if sm.typ == SheetTypeEnum.WorkSheet and sm.visible == SheetVisibleEnum.Visible
        ]

    tables: list[Table] = []
    for sheet_name in sheet_names:
        try:
            sheet = wb.get_sheet_by_name(sheet_name)
        except Exception:
            continue  # Skip sheets that can't be found

        table = _sheet_to_table(sheet, header_row=header_row, max_rows=max_rows)
        tables.append(table)

    # Extract formulas if requested
    if include_formulas:
        _attach_formulas(p, tables, sheet_names)

    # Build metadata
    source = SourceRef(uri=p.as_uri(), mime_type=_XLSX_MIME)
    metadata = DocumentMetadata(
        title=p.stem,
        source=source,
        document_type="xlsx",
        extra={"sheet_count": len(sheet_names)},
    )
    provenance = Provenance(source=source, extractor="kaos-office/xlsx/calamine")

    return TabularDocument(
        metadata=metadata,
        tables=tuple(tables),
        provenance=provenance,
    )


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def list_sheets(path: str | Path) -> list[dict[str, Any]]:
    """List all sheets in an XLSX file with metadata.

    Returns a list of dicts with name, type, visible, dimensions.
    """
    p = Path(path).resolve()
    if not p.is_file():
        msg = f"File not found: {p}"
        raise FileNotFoundError(msg)

    wb = CalamineWorkbook.from_path(str(p))
    result = []
    for sm in wb.sheets_metadata:
        sheet = wb.get_sheet_by_name(sm.name)
        result.append(
            {
                "name": sm.name,
                "type": str(sm.typ),
                "visible": str(sm.visible) == "SheetVisibleEnum.Visible",
                "rows": sheet.height,
                "columns": sheet.width,
            }
        )
    return result


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _sheet_to_table(
    sheet: Any,
    *,
    header_row: int = 0,
    max_rows: int | None = None,
) -> Table:
    """Convert a CalamineSheet to a Table.

    Reads all rows, uses ``header_row`` for column names, infers types
    from the data below the header.
    """
    all_rows = sheet.to_python(skip_empty_area=True)

    if not all_rows:
        return Table(name=sheet.name)

    # Split header and data
    if header_row >= len(all_rows):
        # Header row beyond data — treat all as data with generated headers
        data_rows = all_rows
        n_cols = max(len(r) for r in data_rows) if data_rows else 0
        header = [f"column_{i}" for i in range(n_cols)]
    else:
        header = [
            str(v) if v is not None else f"column_{i}" for i, v in enumerate(all_rows[header_row])
        ]
        data_rows = all_rows[header_row + 1 :]

    if not header:
        return Table(name=sheet.name)

    # Apply max_rows
    total_rows = len(data_rows)
    if max_rows is not None and len(data_rows) > max_rows:
        data_rows = data_rows[:max_rows]

    # Normalize row widths and convert values
    n_cols = len(header)
    normalized_rows: list[tuple[Any, ...]] = []
    for row in data_rows:
        # Pad short rows, truncate long rows
        padded = list(row) + [None] * max(0, n_cols - len(row))
        normalized_rows.append(tuple(padded[:n_cols]))

    # Infer column types from data
    columns: list[Column] = []
    for i, name in enumerate(header):
        col_values = [row[i] for row in normalized_rows if i < len(row)]
        ct = infer_column_type(col_values)
        nullable = any(v is None for v in col_values)
        columns.append(Column(name=name, column_type=ct, nullable=nullable))

    # Build merged cell metadata
    merged = sheet.merged_cell_ranges
    merged_meta: dict[str, Any] = {}
    if merged:
        merged_meta["merged_ranges"] = [f"{_cell_ref(r[0])}: {_cell_ref(r[1])}" for r in merged]

    return Table(
        name=sheet.name,
        columns=tuple(columns),
        rows=tuple(normalized_rows),
        row_count=total_rows,
        metadata=merged_meta,
    )


def _cell_ref(coord: tuple[int, int]) -> str:
    """Convert (row, col) 0-based to Excel-style reference (e.g., A1)."""
    row, col = coord
    letters = ""
    c = col
    while True:
        letters = chr(65 + c % 26) + letters
        c = c // 26 - 1
        if c < 0:
            break
    return f"{letters}{row + 1}"


def _attach_formulas(
    path: Path,
    tables: list[Table],
    sheet_names: list[str],
) -> None:
    """Attach formulas from openpyxl to existing tables.

    openpyxl is much slower than calamine but is the only way to
    read formulas (calamine only returns cached values).
    """
    try:
        import openpyxl
    except ImportError:
        return  # openpyxl not installed — skip formula extraction

    wb = openpyxl.load_workbook(str(path), data_only=False)

    for i, sheet_name in enumerate(sheet_names):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        formulas: dict[str, str] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and cell.value and str(cell.value).startswith("="):
                    coord = cell.coordinate  # e.g., "B5"
                    formulas[coord] = str(cell.value)

        if formulas and i < len(tables):
            # Frozen dataclass — rebuild with updated metadata
            old = tables[i]
            new_meta = dict(old.metadata)
            new_meta["formulas"] = formulas
            tables[i] = Table(
                name=old.name,
                columns=old.columns,
                rows=old.rows,
                row_count=old.row_count,
                metadata=new_meta,
            )

    wb.close()
